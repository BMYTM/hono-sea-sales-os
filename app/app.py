#!/usr/bin/env python3
"""
HONO Proposal Studio — private local backend.

Runs on Boonchoo's Mac (where the base template and client folders live) and
exposes the Proposal Generator wired to the REAL engine: patches the HONO base
template into client-ready .docx files, saves them into client folders, and
lets you browse the archive.

Features:
  • Multi-version generation (several scopes in one client, one click)
  • Excel module-checklist intake (auto-fill countries + module decisions)
  • Optional PDF preview (verify before you send) when LibreOffice is present

Private: binds to 127.0.0.1 only, nothing leaves the machine.

Run:
    python app.py    →    http://127.0.0.1:8765
"""

import io
import os
import json
import shutil
import zipfile
import tempfile
import datetime
import subprocess
import contextlib
from pathlib import Path

from flask import Flask, request, jsonify, send_file

import merge_runs
import patch_proposal

# ── Configuration (override via environment variables) ──────────────────────
HOME = str(Path.home())
DEFAULT_PROJECT = os.path.join(HOME, "Claude", "Projects", "Proposal Generation")
PROJECT_DIR = os.environ.get("HONO_PROJECT_DIR", DEFAULT_PROJECT)
TEMPLATE = os.environ.get(
    "HONO_TEMPLATE",
    os.path.join(PROJECT_DIR, "HONO_HCM Proposal_01072026_SIM_FullHRMSSuite.docx"),
)
# LibreOffice for PDF preview. Common macOS path; override with HONO_SOFFICE.
SOFFICE = os.environ.get("HONO_SOFFICE", "/Applications/LibreOffice.app/Contents/MacOS/soffice")
PORT = int(os.environ.get("HONO_PORT", "8765"))
PREVIEW_DIR = os.path.join(tempfile.gettempdir(), "hono_studio_previews")
os.makedirs(PREVIEW_DIR, exist_ok=True)

# Shared portal libraries (siblings of the Proposal Generation folder by default)
ROOT = os.environ.get("HONO_ROOT", os.path.dirname(PROJECT_DIR))
RFP_DIR = os.environ.get("HONO_RFP_DIR", os.path.join(ROOT, "RFP Library"))
ASSETS_DIR = os.environ.get("HONO_ASSETS_DIR", os.path.join(ROOT, "Sales Assets"))
BANK_PATH = os.path.join(RFP_DIR, "answer_bank.json")
ASSET_CATEGORIES = ["Corporate Decks", "Solution Decks", "Case Studies",
                    "One-Pagers", "Battlecards", "Pricing", "Other"]

def _ensure_dirs():
    for d in (RFP_DIR, os.path.join(RFP_DIR, "Old"), os.path.join(RFP_DIR, "New"), ASSETS_DIR):
        try:
            os.makedirs(d, exist_ok=True)
        except Exception:
            pass
    for c in ASSET_CATEGORIES:
        try:
            os.makedirs(os.path.join(ASSETS_DIR, c), exist_ok=True)
        except Exception:
            pass

# download roots allow-list (path-traversal safe)
def _base_dir(key):
    return {"proposals": PROJECT_DIR, "rfp": RFP_DIR, "assets": ASSETS_DIR}.get(key, PROJECT_DIR)

# Optional team gate. Set HONO_PASSWORD to require a shared login (recommended
# when running on a shared machine or the office network). HONO_USER defaults
# to "team". Leave HONO_PASSWORD unset for open localhost-only use.
AUTH_USER = os.environ.get("HONO_USER", "team")
AUTH_PASS = os.environ.get("HONO_PASSWORD", "")
# Bind host: 127.0.0.1 (this machine only) unless HONO_HOST is set to 0.0.0.0
# so teammates on the LAN can reach it.
BIND_HOST = os.environ.get("HONO_HOST", "127.0.0.1")

app = Flask(__name__)


@app.before_request
def _guard():
    if not AUTH_PASS:
        return  # no password configured -> open (localhost use)
    from flask import request as _rq, Response as _Resp
    auth = _rq.authorization
    if not auth or auth.username != AUTH_USER or auth.password != AUTH_PASS:
        return _Resp("Authentication required.", 401,
                     {"WWW-Authenticate": 'Basic realm="HONO Sales OS"'})


# ── Helpers ─────────────────────────────────────────────────────────────────
def _month_num(name):
    months = ["january", "february", "march", "april", "may", "june", "july",
              "august", "september", "october", "november", "december"]
    return months.index(name.strip().lower()) + 1


def soffice_path():
    if os.path.exists(SOFFICE):
        return SOFFICE
    found = shutil.which("soffice") or shutil.which("libreoffice")
    return found


def safe_within(base, target):
    base = os.path.realpath(base)
    target = os.path.realpath(target)
    return target == base or target.startswith(base + os.sep)


def build_config(client, version, unpacked_dir):
    """Translate shared client info + one version into patch_proposal's config."""
    countries = [{"name": c["name"], "count": int(c["count"])}
                 for c in version["countries"]
                 if c.get("name") and str(c.get("count", "")).strip()]
    total = sum(c["count"] for c in countries)
    d = client["date"]
    return {
        "client_name": client["client_name"],
        "client_short": client["client_short"],
        "client_code": client["client_code"],
        "contact_name": client["contact_name"],
        "proposal_date": {"day": int(d["day"]), "suffix": client.get("suffix", ""),
                          "month": d["month"], "year": int(d["year"])},
        "scope": version["scope"],
        "headcount": {"total": total, "countries": countries},
        "pricing": {"pepm": float(version["pepm"]), "currency": version["currency"]},
        "modules": {
            "remove": version.get("remove", []),
            "defer_attendance": bool(version.get("defer_attendance", False)),
            "payroll_exclude_countries": version.get("payroll_exclude", []),
            "phase2_lms": bool(version.get("phase2_lms", True)),
        },
        "unpacked_dir": unpacked_dir,
    }, total


def make_docx(client, version):
    """Run the full pipeline for a single version. Returns a result dict."""
    work = tempfile.mkdtemp(prefix="hono_studio_")
    try:
        draft = os.path.join(work, "draft.docx")
        shutil.copy(TEMPLATE, draft)
        unpacked = os.path.join(work, "unpacked")
        with zipfile.ZipFile(draft) as zf:
            zf.extractall(unpacked)
        for root, _, files in os.walk(unpacked):
            for f in files:
                p = os.path.join(root, f)
                if os.path.islink(p):
                    os.unlink(p)

        merge_runs.merge_document(os.path.join(unpacked, "word", "document.xml"))

        config, total = build_config(client, version, unpacked)
        cfg_path = os.path.join(work, "config.json")
        with open(cfg_path, "w", encoding="utf-8") as f:
            json.dump(config, f)

        buf = io.StringIO()
        with contextlib.redirect_stdout(buf):
            clean = patch_proposal.patch(cfg_path)
        report = buf.getvalue()

        docxml = open(os.path.join(unpacked, "word", "document.xml"), encoding="utf-8").read()
        headless_ok = "Headless API" in docxml
        if not headless_ok:
            return {"ok": False, "scope": version["scope"], "report": report,
                    "error": "CRITICAL: 'Headless API & Token Usage' clause missing — not saved."}

        out_docx = os.path.join(work, "output.docx")
        with zipfile.ZipFile(out_docx, "w", zipfile.ZIP_DEFLATED) as zf:
            for root, _, files in os.walk(unpacked):
                for f in files:
                    fp = os.path.join(root, f)
                    zf.write(fp, os.path.relpath(fp, unpacked))

        client_folder = os.path.join(PROJECT_DIR, client["client_short"])
        os.makedirs(client_folder, exist_ok=True)
        d = config["proposal_date"]
        stamp = f"{d['day']:02d}{_month_num(d['month']):02d}{d['year']}"
        fname = f"HONO_HCM Proposal_{stamp}_{client['client_short']}_{version['scope']}.docx"
        final_path = os.path.join(client_folder, fname)
        shutil.copy(out_docx, final_path)

        # auto-generate a PDF next to the .docx when LibreOffice is available
        pdf_rel = None
        so = soffice_path()
        if so and os.environ.get("HONO_AUTO_PDF", "1") != "0":
            try:
                subprocess.run([so, "--headless", "--convert-to", "pdf", "--outdir", client_folder, final_path],
                               check=True, timeout=180, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
                pdf_path = os.path.splitext(final_path)[0] + ".pdf"
                if os.path.isfile(pdf_path):
                    pdf_rel = os.path.relpath(pdf_path, PROJECT_DIR)
            except Exception:
                pass

        pepm = float(version["pepm"]); annual = int(total * pepm * 12); impl = round(annual * 0.75)
        proposal_id = f"{client['client_code']}-{version['scope'].upper()}-{stamp}"

        return {
            "ok": True, "clean": clean, "report": report, "headless_ok": headless_ok,
            "scope": version["scope"], "file": fname,
            "rel": os.path.relpath(final_path, PROJECT_DIR),
            "pdf_rel": pdf_rel,
            "client_folder": client_folder,
            "summary": {"total": total, "pepm": pepm, "currency": version["currency"],
                        "annual": annual, "impl": impl, "proposal_id": proposal_id},
        }
    finally:
        shutil.rmtree(work, ignore_errors=True)


# ── Routes ──────────────────────────────────────────────────────────────────
@app.get("/")
def index():
    return send_file(os.path.join(os.path.dirname(__file__), "dashboard.html"))


@app.get("/studio")
def studio():
    return send_file(os.path.join(os.path.dirname(__file__), "studio.html"))


@app.get("/api/env")
def env():
    return jsonify({
        "project_dir": PROJECT_DIR, "template": TEMPLATE,
        "template_exists": os.path.exists(TEMPLATE),
        "project_exists": os.path.isdir(PROJECT_DIR),
        "soffice_available": bool(soffice_path()),
    })


@app.post("/api/generate")
def generate():
    payload = request.get_json(force=True)
    if not os.path.exists(TEMPLATE):
        return jsonify({"ok": False, "error": f"Base template not found at:\n{TEMPLATE}\nSet HONO_TEMPLATE."}), 400
    client = payload["client"]
    versions = payload["versions"]
    results = [make_docx(client, v) for v in versions]
    client_folder = os.path.join(PROJECT_DIR, client["client_short"])
    return jsonify({"ok": any(r.get("ok") for r in results),
                    "client_folder": client_folder, "results": results})


@app.post("/api/parse-checklist")
def parse_checklist():
    """Parse a client's module-checklist xlsx into countries + module decisions."""
    try:
        import openpyxl
    except ImportError:
        return jsonify({"error": "openpyxl not installed. Run: pip install openpyxl"}), 500
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400
    f = request.files["file"]
    tmp = os.path.join(tempfile.gettempdir(), "hono_checklist.xlsx")
    f.save(tmp)
    try:
        wb = openpyxl.load_workbook(tmp, data_only=True)
    except Exception as e:
        return jsonify({"error": f"Could not read workbook: {e}"}), 400

    sheets = wb.sheetnames
    countries, modules = [], []

    # Sheet 2 = Countries & Headcount (col B country, col C headcount)
    if len(sheets) >= 2:
        ws = wb[sheets[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 3:
                continue
            name, count = row[1], row[2]
            if name and isinstance(count, (int, float)):
                countries.append({"name": str(name).strip(), "count": int(count)})

    # Sheet 3 = Module Checklist (col B module, col C Include?, col D Priority)
    defer_attendance = False
    remove = []
    if len(sheets) >= 3:
        ws = wb[sheets[2]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 3:
                continue
            mod, include, priority = row[1], row[2], (row[3] if len(row) > 3 else None)
            if not mod:
                continue
            mod = str(mod).strip()
            inc = str(include).strip().lower() if include is not None else ""
            pri = str(priority).strip().lower() if priority is not None else ""
            decision = "include"
            if inc in ("no", "n", "false", "0", "exclude"):
                decision = "exclude"; remove.append(mod)
            elif pri in ("maybe", "future"):
                decision = "defer"
            elif pri == "phase 2":
                decision = "phase2"
            modules.append({"module": mod, "decision": decision})
            if "attendance" in mod.lower() and decision in ("defer", "exclude"):
                defer_attendance = True

    return jsonify({"sheets": sheets, "countries": countries, "modules": modules,
                    "remove": remove, "defer_attendance": defer_attendance})


@app.get("/api/proposals")
def proposals():
    if not os.path.isdir(PROJECT_DIR):
        return jsonify({"clients": [], "error": f"Project folder not found: {PROJECT_DIR}"})
    clients = []
    for entry in sorted(os.listdir(PROJECT_DIR)):
        folder = os.path.join(PROJECT_DIR, entry)
        if not os.path.isdir(folder):
            continue
        files = []
        for fn in sorted(os.listdir(folder)):
            fp = os.path.join(folder, fn)
            if os.path.isfile(fp) and fn.lower().endswith((".docx", ".pdf")):
                st = os.stat(fp)
                files.append({"name": fn, "rel": os.path.relpath(fp, PROJECT_DIR),
                              "size_kb": round(st.st_size / 1024),
                              "modified": datetime.datetime.fromtimestamp(st.st_mtime).strftime("%Y-%m-%d %H:%M")})
        if files:
            clients.append({"client": entry, "count": len(files), "files": files})
    return jsonify({"clients": clients, "project_dir": PROJECT_DIR})


@app.get("/api/download")
def download():
    base = _base_dir(request.args.get("base", "proposals"))
    rel = request.args.get("path", "")
    target = os.path.join(base, rel)
    if not safe_within(base, target) or not os.path.isfile(target):
        return jsonify({"error": "Not found"}), 404
    return send_file(target, as_attachment=True)


@app.get("/api/preview")
def preview():
    """Convert a saved .docx to PDF (LibreOffice) and return it inline for review."""
    so = soffice_path()
    if not so:
        return jsonify({"error": "LibreOffice not found — PDF preview unavailable."}), 400
    rel = request.args.get("path", "")
    src = os.path.join(PROJECT_DIR, rel)
    if not safe_within(PROJECT_DIR, src) or not os.path.isfile(src):
        return jsonify({"error": "Not found"}), 404
    try:
        subprocess.run([so, "--headless", "--convert-to", "pdf", "--outdir", PREVIEW_DIR, src],
                       check=True, timeout=120, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
    except Exception as e:
        return jsonify({"error": f"PDF conversion failed: {e}"}), 500
    pdf = os.path.join(PREVIEW_DIR, os.path.splitext(os.path.basename(src))[0] + ".pdf")
    if not os.path.isfile(pdf):
        return jsonify({"error": "PDF not produced"}), 500
    return send_file(pdf, mimetype="application/pdf")


# ── RFP Library: pages, answer bank, auto-fill ──────────────────────────────
import re as _re
import math as _math
from collections import Counter as _Counter

_STOP = set("the a an of to and or for in on at is are be with as by from your our "
            "you we do does can will please provide describe list what how which".split())


def _tokens(s):
    return [t for t in _re.findall(r"[a-z0-9]+", str(s).lower()) if t not in _STOP and len(t) > 1]


def _cosine(a, b):
    ca, cb = _Counter(a), _Counter(b)
    common = set(ca) & set(cb)
    num = sum(ca[t] * cb[t] for t in common)
    da = _math.sqrt(sum(v * v for v in ca.values()))
    db = _math.sqrt(sum(v * v for v in cb.values()))
    return num / (da * db) if da and db else 0.0


def _load_bank():
    if os.path.isfile(BANK_PATH):
        try:
            return json.load(open(BANK_PATH, encoding="utf-8"))
        except Exception:
            return []
    return []


def _save_bank(bank):
    _ensure_dirs()
    with open(BANK_PATH, "w", encoding="utf-8") as f:
        json.dump(bank, f, ensure_ascii=False, indent=1)


def _best_answer(question, bank):
    qt = _tokens(question)
    best, score = None, 0.0
    for e in bank:
        s = _cosine(qt, _tokens(e["question"]))
        if s > score:
            best, score = e, s
    return best, round(score, 3)


@app.get("/rfp")
def rfp_page():
    return send_file(os.path.join(os.path.dirname(__file__), "rfp.html"))


@app.get("/api/rfp/summary")
def rfp_summary():
    _ensure_dirs()
    bank = _load_bank()
    def _count(sub):
        d = os.path.join(RFP_DIR, sub)
        return len([f for f in os.listdir(d)]) if os.path.isdir(d) else 0
    tags = sorted({t for e in bank for t in e.get("tags", [])})
    return jsonify({"bank_count": len(bank), "old_count": _count("Old"),
                    "new_count": _count("New"), "tags": tags, "rfp_dir": RFP_DIR})


@app.post("/api/rfp/import-bank")
def rfp_import_bank():
    """Import an old RFP (xlsx: question column + answer column) into the answer bank."""
    try:
        import openpyxl
    except ImportError:
        return jsonify({"error": "openpyxl not installed."}), 500
    if "file" not in request.files:
        return jsonify({"error": "No file"}), 400
    f = request.files["file"]
    tag = (request.form.get("tag") or "").strip()
    _ensure_dirs()
    saved = os.path.join(RFP_DIR, "Old", f.filename)
    f.save(saved)
    try:
        wb = openpyxl.load_workbook(saved, data_only=True)
    except Exception as e:
        return jsonify({"error": f"Could not read: {e}"}), 400
    ws = wb[wb.sheetnames[0]]
    # detect question/answer columns from header row, else assume col1=Q col2=A
    header = [str(c.value).lower() if c.value else "" for c in ws[1]]
    qcol = next((i for i, h in enumerate(header) if "question" in h or "requirement" in h), 0)
    acol = next((i for i, h in enumerate(header) if "answer" in h or "response" in h or "reply" in h), 1)
    bank = _load_bank()
    existing = {e["question"].strip().lower() for e in bank}
    added = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) <= max(qcol, acol):
            continue
        q, a = row[qcol], row[acol]
        if q and a and str(q).strip().lower() not in existing:
            bank.append({"question": str(q).strip(), "answer": str(a).strip(),
                         "source": f.filename, "tags": [tag] if tag else []})
            existing.add(str(q).strip().lower())
            added += 1
    _save_bank(bank)
    return jsonify({"ok": True, "added": added, "bank_count": len(bank)})


@app.post("/api/rfp/autofill")
def rfp_autofill():
    """Upload a new RFP (xlsx/csv with a question column). Auto-fill from the bank."""
    try:
        import openpyxl
    except ImportError:
        return jsonify({"error": "openpyxl not installed."}), 500
    if "file" not in request.files:
        return jsonify({"error": "No file"}), 400
    f = request.files["file"]
    threshold = float(request.form.get("threshold", "0.35"))
    _ensure_dirs()
    src = os.path.join(RFP_DIR, "New", f.filename)
    f.save(src)
    bank = _load_bank()
    if not bank:
        return jsonify({"error": "Answer bank is empty — import an old RFP first."}), 400

    # read questions (xlsx first sheet or csv)
    questions = []
    if f.filename.lower().endswith(".csv"):
        import csv
        with open(src, encoding="utf-8", errors="ignore") as fh:
            for i, row in enumerate(csv.reader(fh)):
                if i == 0 and row and "question" in row[0].lower():
                    continue
                if row and row[0].strip():
                    questions.append(row[0].strip())
    else:
        wb = openpyxl.load_workbook(src, data_only=True)
        ws = wb[wb.sheetnames[0]]
        header = [str(c.value).lower() if c.value else "" for c in ws[1]]
        qcol = next((i for i, h in enumerate(header) if "question" in h or "requirement" in h), 0)
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and len(row) > qcol and row[qcol]:
                questions.append(str(row[qcol]).strip())

    rows, filled, review = [], 0, 0
    for q in questions:
        best, score = _best_answer(q, bank)
        if best and score >= threshold:
            rows.append({"question": q, "answer": best["answer"], "confidence": score,
                         "source": best["source"], "status": "filled"})
            filled += 1
        else:
            rows.append({"question": q, "answer": best["answer"] if best else "",
                         "confidence": score, "source": best["source"] if best else "",
                         "status": "review"})
            review += 1

    # write a filled workbook
    out = openpyxl.Workbook()
    ws = out.active; ws.title = "Auto-filled RFP"
    ws.append(["Question", "Suggested Answer", "Confidence", "Source", "Status"])
    for r in rows:
        ws.append([r["question"], r["answer"], r["confidence"], r["source"], r["status"]])
    base = os.path.splitext(f.filename)[0]
    out_name = f"{base}_autofilled.xlsx"
    out_path = os.path.join(RFP_DIR, "New", out_name)
    out.save(out_path)

    return jsonify({"ok": True, "total": len(rows), "filled": filled, "review": review,
                    "rows": rows[:200], "file": out_name,
                    "rel": os.path.relpath(out_path, RFP_DIR)})


@app.get("/api/rfp/list")
def rfp_list():
    _ensure_dirs()
    def _files(sub):
        d = os.path.join(RFP_DIR, sub); out = []
        if os.path.isdir(d):
            for fn in sorted(os.listdir(d)):
                fp = os.path.join(d, fn)
                if os.path.isfile(fp):
                    st = os.stat(fp)
                    out.append({"name": fn, "rel": os.path.relpath(fp, RFP_DIR),
                                "size_kb": round(st.st_size / 1024),
                                "modified": datetime.datetime.fromtimestamp(st.st_mtime).strftime("%Y-%m-%d %H:%M")})
        return out
    return jsonify({"old": _files("Old"), "new": _files("New")})


# ── Sales Assets: decks & collateral ────────────────────────────────────────
@app.get("/decks")
def decks_page():
    return send_file(os.path.join(os.path.dirname(__file__), "decks.html"))


@app.get("/api/assets/list")
def assets_list():
    _ensure_dirs()
    cats = []
    for c in ASSET_CATEGORIES:
        d = os.path.join(ASSETS_DIR, c); files = []
        if os.path.isdir(d):
            for fn in sorted(os.listdir(d)):
                fp = os.path.join(d, fn)
                if os.path.isfile(fp):
                    st = os.stat(fp)
                    files.append({"name": fn, "rel": os.path.relpath(fp, ASSETS_DIR),
                                  "size_kb": round(st.st_size / 1024),
                                  "modified": datetime.datetime.fromtimestamp(st.st_mtime).strftime("%Y-%m-%d %H:%M")})
        cats.append({"category": c, "count": len(files), "files": files})
    return jsonify({"categories": cats, "assets_dir": ASSETS_DIR})


@app.post("/api/assets/upload")
def assets_upload():
    _ensure_dirs()
    if "file" not in request.files:
        return jsonify({"error": "No file"}), 400
    cat = request.form.get("category", "Other")
    if cat not in ASSET_CATEGORIES:
        cat = "Other"
    f = request.files["file"]
    dest = os.path.join(ASSETS_DIR, cat, f.filename)
    f.save(dest)
    return jsonify({"ok": True, "category": cat, "name": f.filename})


if __name__ == "__main__":
    _ensure_dirs()
    reach = "127.0.0.1" if BIND_HOST == "127.0.0.1" else "<this machine's IP>"
    print("=" * 60)
    print(" HONO Sales OS (private)")
    print(f"   template : {TEMPLATE}")
    print(f"   archive  : {PROJECT_DIR}")
    print(f"   PDF      : {'yes' if soffice_path() else 'no (install LibreOffice)'}")
    print(f"   auth     : {'ON (shared password)' if AUTH_PASS else 'OFF (localhost only)'}")
    print(f"   dashboard: http://{reach}:{PORT}/")
    print(f"   studio   : http://{reach}:{PORT}/studio")
    print("=" * 60)
    app.run(host=BIND_HOST, port=PORT, debug=False)
